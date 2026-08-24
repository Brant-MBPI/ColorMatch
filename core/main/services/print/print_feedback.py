import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import Q
from django.http import HttpResponse
from datetime import datetime, date
from main.models import (
    tbl_feedback_details, tbl_cmf, tbl_rs, tbl_cmf_dates, 
    tbl_cmf_formula, tbl_cmf_pending_completed, tbl_resins_selected,
    tbl_cmf_process02, tbl_cmf_color_req, tbl_mb_extruder_formula,
    tbl_dc_extruder_formula
)

def generate_feedback_excel(date_from=None, date_to=None):
    """
    Generates Excel for Feedback Records.
    Exclusively uses tbl_cmf_dates for sorting and filtering.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback Records"

    # 1. Define Headers
    headers = [
        "Matching No.", "Customer", "Date Created", "Date Lab Received", "Target Date",
        "Primary Color", "Color Description", "End Product", "Matching Type", "Salesman",
        "Color Req.", "Resin", "Process", "Type of Colorant", "Date Given Sample",
        "Set/PC", "Quantity Given", "Code Submitted", "Dosage", "Lot #", "AR #",
        "Date Standard & Result", "Comment", "Storage Details"
    ]

    # 2. Styles
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 3. Fetch Base Feedback Records
    feedback_qs = tbl_feedback_details.objects.select_related('cm_no', 'rs_no').all()
    
    rows_to_sort = []

    for fb in feedback_qs:
        is_cmf = True if fb.cm_no else False
        parent = fb.cm_no if is_cmf else fb.rs_no
        if not parent:
            continue

        # FETCH DATE FROM THE SINGLE SOURCE OF TRUTH
        dates_rec = tbl_cmf_dates.objects.filter(
            Q(cm_no=parent) if is_cmf else Q(rs_no=parent)
        ).first()

        created_date = dates_rec.form_made if dates_rec else None

        # Filter by Date Range if provided
        if date_from and date_to:
            if not created_date or not (date_from <= created_date <= date_to):
                continue

        # FETCH RELATED INFO
        pending = tbl_cmf_pending_completed.objects.filter(
            Q(cm_no=parent) if is_cmf else Q(rs_no=parent)
        ).first()

        # Specific Color Req Logic for both CMF and RS
        color_req = tbl_cmf_color_req.objects.filter(
            Q(cm_no=parent) if is_cmf else Q(rs_no=parent)
        ).first()
        color_req_val = color_req.name if color_req else ""

        if is_cmf:
            formula = tbl_cmf_formula.objects.filter(cm_no=parent).first()
            customer = formula.customer if formula else ""
            end_product = formula.finished_product if formula else ""
            dosage = formula.dosage if formula else ""
            salesman = parent.sm.name if parent.sm else ""
            primary_color = parent.in_code_no.color if parent.in_code_no else ""
            process_list = tbl_cmf_process02.objects.filter(cmf_formula_no=formula).values_list('process_no__name', flat=True) if formula else []
        else:
            customer = parent.customer or ""
            end_product = parent.finished_product or ""
            dosage = parent.dosage or ""
            salesman = parent.sm_no.name if parent.sm_no else ""
            primary_color = parent.primary_color or ""
            process_list = tbl_cmf_process02.objects.filter(rs_no=parent).values_list('process_no__name', flat=True)

        resins = ", ".join(tbl_resins_selected.objects.filter(
            cm_no=parent if is_cmf else None, 
            rs_no=parent if not is_cmf else None
        ).values_list('resin_no__abbreviation', flat=True))
        
        processes = ", ".join(process_list)

        # Final Code & Lot Info
        code_sub = ""
        lot_no = "None"
        final_f = tbl_mb_extruder_formula.objects.filter(
            Q(cm_no=parent) if is_cmf else Q(rs_no=parent), is_final=True
        ).select_related('code').first()
        
        if final_f:
            code_sub = final_f.code.product_code if final_f.code else ""
            lot_no = final_f.lot_no or "None"
        else:
            final_dc = tbl_dc_extruder_formula.objects.filter(
                Q(cm_no=parent) if is_cmf else Q(rs_no=parent), is_final=True
            ).select_related('code').first()
            if final_dc:
                code_sub = final_dc.code.product_code if final_dc.code else ""
                lot_no = "None"

        # Build Data Row
        data_row = [
            parent.cm_no if is_cmf else parent.rs_no,
            customer,
            created_date.strftime('%m/%d/%Y') if created_date else "",
            dates_rec.date_received_lab if dates_rec else "",
            dates_rec.due_date_lab.strftime('%m/%d/%Y') if dates_rec and dates_rec.due_date_lab else "",
            primary_color,
            parent.color_desc or "",
            end_product,
            parent.matching_type or "",
            salesman,
            color_req_val,
            resins,
            processes,
            parent.colorant_type or "",
            pending.date_submitted.strftime('%m/%d/%Y') if pending and pending.date_submitted else "",
            fb.pieces or "",
            fb.quantity_given or "",
            code_sub,
            dosage,
            lot_no,
            pending.ar_no if pending else "",
            fb.date_sample_received.strftime('%m/%d/%Y') if fb.date_sample_received else "",
            fb.comment or "",
            fb.storage_details or ""
        ]

        # Use 1900-01-01 as proxy for missing dates to keep them at the top
        rows_to_sort.append((created_date or date(1900, 1, 1), data_row))

    # 4. Sort Ascending (Oldest at top, Latest at bottom)
    rows_to_sort.sort(key=lambda x: x[0])

    # 5. Write to Sheet
    for idx, (dt, row_data) in enumerate(rows_to_sort, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=idx, column=col_idx, value=value)

    # 6. Auto-fit Columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    return wb, rows_to_sort

def export_feedback_excel(request):
    from_date_str = request.GET.get('from')
    to_date_str = request.GET.get('to')
    
    date_from = None
    date_to = None
    
    try:
        if from_date_str: date_from = datetime.strptime(from_date_str, '%m/%d/%Y').date()
        if to_date_str: date_to = datetime.strptime(to_date_str, '%m/%d/%Y').date()
    except ValueError: pass

    wb, sorted_data = generate_feedback_excel(date_from, date_to)
    
    # 7. Filename Construction (MMDDYY-MMDDYY)
    if date_from and date_to:
        f_range = f"{date_from:%m%d%y}-{date_to:%m%d%y}"
    elif sorted_data:
        # Get actual range from sorted data
        first_date = sorted_data[0][0].strftime('%m%d%y')
        last_date = sorted_data[-1][0].strftime('%m%d%y')
        f_range = f"{first_date}-{last_date}"
    else:
        f_range = "Empty"

    filename = f"Feedback_Report_{f_range}.xlsx"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response