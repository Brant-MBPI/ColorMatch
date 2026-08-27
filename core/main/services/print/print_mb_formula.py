import os
import io
import tempfile
import threading
import uuid
import subprocess
import shutil
from decimal import Decimal
from datetime import datetime

# Cross-platform Excel & PDF manipulation
import openpyxl
from openpyxl.worksheet.page import PageMargins
import pymupdf 

from django.contrib import messages
from django.http import HttpResponse, HttpResponseServerError, JsonResponse
from django.shortcuts import redirect
from django.views.decorators.clickjacking import xframe_options_exempt

from main.utils.log_audit_trail import log_audit
from main.models import (
    tbl_mb_extruder_formula, tbl_mb_extruder_formula02,
    tbl_cmf_formula, tbl_resins_selected,
)

# Lock ensures LibreOffice instances don't collide
_excel_lock = threading.Lock()

MB_TEMPLATE_PATH = os.path.join('main', 'templates', 'print_excel', 'mb_formula_template.xlsx')

# Target Dimensions (Half-Letter Landscape)
MB_PDF_WIDTH_IN = 8.5
MB_PDF_HEIGHT_IN = 6.5

# Material row settings
MATERIAL_START_ROW = 13
MATERIAL_MAX_ROWS = 10

# Excel number formats
FMT_PERCENT_4DP = '0.0000'
FMT_WEIGHT_7DP_G = '0.0000000"g"'
FMT_DOSAGE_PCT = '0.00"%"'

def _get_libreoffice_executable():
    """Finds the LibreOffice/soffice executable based on OS."""
    if os.name == 'nt':  # Windows
        paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for path in paths:
            if os.path.exists(path): return path
        return "soffice"
    return "libreoffice" # Linux/Synology

def _resize_pdf_to_fixed_size(input_pdf_path, output_pdf_path, width_in=8.5, height_in=6.5):
    """
    Rescales the PDF. Uses a robust method compatible with older PyMuPDF versions.
    It takes the top portion of the raw page and fits it to the 8.5x6.5 canvas.
    """
    target_w_pt = width_in * 72
    target_h_pt = height_in * 72

    src = pymupdf.open(input_pdf_path)
    dst = pymupdf.open()
    
    if len(src) > 0:
        page = src[0]
        
        # Calculate the area to capture. 
        # Standard Letter is 612pt wide. Your form (A-G) usually fits this width.
        # We take the top 468pts (6.5 inches) of the source page.
        crop_rect = pymupdf.Rect(0, 0, page.rect.width, 468) 
        
        new_page = dst.new_page(width=target_w_pt, height=target_h_pt)
        
        # Place the content. 'clip' acts as the crop tool.
        new_page.show_pdf_page(
            pymupdf.Rect(0, 0, target_w_pt, target_h_pt), 
            src, 
            0,
            clip=crop_rect
        )
    
    dst.save(output_pdf_path)
    dst.close()
    src.close()

def _fill_and_export_mb_formula_via_excel(template_abs_path, pdf_path, data):
    """Fills Excel and prepares it for PDF conversion."""
    header = data['header']
    ingredients = data['ingredients']

    try:
        wb = openpyxl.load_workbook(template_abs_path, data_only=False, keep_vba=False)
        ws = wb.worksheets[0]
        ws.views.sheetView = [] # Fix for NoneType crash
    except Exception as e:
        raise RuntimeError(f"OpenPyXL Load Error: {str(e)}")

    # Set Print Area strictly to your form area (A1 to G25)
    try:
        ws.page_margins = PageMargins(left=0.1, right=0.1, top=0.1, bottom=0.1, header=0, footer=0)
        ws.print_area = 'A1:G25' 
        if hasattr(ws, 'page_setup') and ws.page_setup:
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 1
            ws.page_setup.fitToWidth = 1
            ws.page_setup.horizontalCentered = True
            ws.page_setup.verticalCentered = True
    except:
        pass

    def set_cell(addr, value, number_format=None):
        try:
            ws[addr] = value
            if number_format: ws[addr].number_format = number_format
        except: pass

    # --- FILLING DATA ---
    set_cell('C6', header.date.strftime('%m/%d/%Y') if header.date else "")
    set_cell('C7', header.code.product_code if header.code else "")
    set_cell('C8', data['customer'])
    set_cell('C9', header.lot_no)
    set_cell('C10', data['color'])
    set_cell('G6', data['parent_no'])
    set_cell('G7', data['resin'])
    set_cell('G8', _to_num(data['dosage']), number_format=FMT_DOSAGE_PCT)
    set_cell('G9', header.mixing_time)
    set_cell('G10', data['application'])

    for i in range(MATERIAL_MAX_ROWS):
        row_num = MATERIAL_START_ROW + i
        if i < len(ingredients):
            ing = ingredients[i]
            set_cell(f'A{row_num}', ing.material)
            set_cell(f'C{row_num}', _to_num(ing.value), number_format=FMT_PERCENT_4DP)
            set_cell(f'G{row_num}', _to_num(ing.weight), number_format=FMT_WEIGHT_7DP_G)
        else:
            set_cell(f'A{row_num}', "")

    set_cell('C23', _to_num(sum((Decimal(ing.value or 0) for ing in ingredients), Decimal('0'))), number_format=FMT_PERCENT_4DP)
    set_cell('G23', _to_num(header.total_weight), number_format=FMT_WEIGHT_7DP_G)
    set_cell('B24', header.matched_by); set_cell('B25', header.weighted_by)
    set_cell('D24', header.notes); set_cell('G24', header.encoded_by)

    # Save Excel
    temp_dir = os.path.dirname(pdf_path)
    temp_xlsx = os.path.join(temp_dir, f"temp_{uuid.uuid4().hex}.xlsx")
    try:
        wb.save(temp_xlsx)
    except:
        try:
            ws.page_setup = None 
            wb.save(temp_xlsx)
        except Exception as e:
            raise RuntimeError(f"Excel Save failure: {str(e)}")

    # Convert to PDF
    profile_path = os.path.join(temp_dir, f"lo_profile_{uuid.uuid4().hex}")
    os.makedirs(profile_path, exist_ok=True)
    try:
        libreoffice_bin = _get_libreoffice_executable()
        profile_url = f"file:///{profile_path.replace(os.sep, '/')}"
        command = [
            f'"{libreoffice_bin}"', '--headless', f'"-env:UserInstallation={profile_url}"',
            '--convert-to pdf', f'--outdir "{temp_dir}"', f'"{temp_xlsx}"'
        ]
        subprocess.run(" ".join(command), shell=True, check=True, capture_output=True, timeout=30)
        generated_pdf = temp_xlsx.replace('.xlsx', '.pdf')
        if os.path.exists(generated_pdf):
            if os.path.exists(pdf_path): os.remove(pdf_path)
            os.rename(generated_pdf, pdf_path)
    finally:
        if os.path.exists(temp_xlsx): os.remove(temp_xlsx)
        shutil.rmtree(profile_path, ignore_errors=True)

def _fetch_mb_formula_data(formula_id):
    header = tbl_mb_extruder_formula.objects.select_related('cm_no', 'rs_no', 'code').get(pk=formula_id)
    ingredients = list(tbl_mb_extruder_formula02.objects.filter(mb=header).order_by('id')[:MATERIAL_MAX_ROWS])
    customer, color, resin, application, dosage, parent_no = "", "", "", "", "", ""
    if header.cm_no:
        parent_no, color = header.cm_no.cm_no, header.cm_no.color_desc
        formula_info = tbl_cmf_formula.objects.filter(cm_no=header.cm_no).first()
        if formula_info:
            customer, application, dosage = formula_info.customer, formula_info.finished_product, formula_info.dosage
        resin = ", ".join(tbl_resins_selected.objects.filter(cm_no=header.cm_no).values_list('resin_no__abbreviation', flat=True))
    elif header.rs_no:
        parent_no, customer, color, application = header.rs_no.rs_no, header.rs_no.customer, header.rs_no.color_desc, header.rs_no.finished_product
        dosage = getattr(header.rs_no, 'dosage', '')
        resin = ", ".join(tbl_resins_selected.objects.filter(rs_no=header.rs_no).values_list('resin_no__abbreviation', flat=True))
    return {
        'header': header, 'ingredients': ingredients, 'customer': customer,
        'color': color, 'resin': resin, 'application': application,
        'dosage': dosage, 'parent_no': parent_no,
    }

def _to_num(val):
    if val is None or val == "": return 0
    try: return float(val)
    except: return 0

@xframe_options_exempt
def print_mb_formula_preview(request, formula_id):
    try:
        data = _fetch_mb_formula_data(formula_id)
    except Exception as e:
        return HttpResponseServerError(f"Data error: {str(e)}")

    template_abs_path = os.path.abspath(MB_TEMPLATE_PATH)
    with _excel_lock:
        with tempfile.TemporaryDirectory() as tmpdir:
            raw_pdf_path = os.path.join(tmpdir, "raw.pdf")
            final_pdf_path = os.path.join(tmpdir, "final.pdf")
            try:
                _fill_and_export_mb_formula_via_excel(template_abs_path, raw_pdf_path, data)
                # Resizing call now uses the robust crop-and-fit method
                _resize_pdf_to_fixed_size(raw_pdf_path, final_pdf_path, width_in=MB_PDF_WIDTH_IN, height_in=MB_PDF_HEIGHT_IN)
                with open(final_pdf_path, 'rb') as f:
                    pdf_bytes = f.read()
            except Exception as e:
                return HttpResponseServerError(f"PDF export failed: {str(e)}")

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'inline'
    response['X-Frame-Options'] = 'SAMEORIGIN'
    return response

def log_formula_print(request, formula_id):
    try:
        formula = tbl_mb_extruder_formula.objects.get(pk=formula_id)
        desc = f"Printed MB Formula (Lot: {formula.lot_no or 'N/A'})"
        log_audit(request, "Printed", desc)
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)