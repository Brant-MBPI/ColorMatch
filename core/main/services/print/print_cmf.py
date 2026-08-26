import os
import io
import tempfile
import threading
import uuid
import subprocess
import shutil
from datetime import datetime

# Cross-platform Excel & PDF manipulation
import openpyxl
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment 
import pymupdf 

from django.contrib import messages
from django.http import HttpResponse, HttpResponseServerError, JsonResponse
from django.shortcuts import redirect
from django.views.decorators.clickjacking import xframe_options_exempt

from main.utils.log_audit_trail import log_audit
from main.models import (
    tbl_cmf, tbl_cmf_dates, tbl_cmf_formula, tbl_cmf_color_req,
    tbl_resins_selected, tbl_cmf_process02, tbl_cmf_specification02,
    tbl_mb_extruder_formula, tbl_dc_extruder_formula
)

# Lock ensures LibreOffice instances don't collide
_excel_lock = threading.Lock()

TEMPLATE_PATH = os.path.join('main', 'templates', 'print_excel', 'new_cmf_template.xlsx')

def _get_libreoffice_executable():
    """Finds the LibreOffice/soffice executable based on OS."""
    if os.name == 'nt':  # Windows
        paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for path in paths:
            if os.path.exists(path):
                return path
        return "soffice"
    return "libreoffice" # Linux/Synology

def _resize_pdf_to_fixed_size(input_pdf_path, output_pdf_path, width_in=6.5, height_in=8.5):
    """Rescales every page using pymupdf to exactly 6.5 x 8.5 inches."""
    target_w_pt = width_in * 72
    target_h_pt = height_in * 72

    src = pymupdf.open(input_pdf_path)
    dst = pymupdf.open()

    for page in src:
        new_page = dst.new_page(width=target_w_pt, height=target_h_pt)
        # Scale content to fill the target size exactly
        new_page.show_pdf_page(
            pymupdf.Rect(0, 0, target_w_pt, target_h_pt),
            src,
            page.number,
        )

    dst.save(output_pdf_path)
    dst.close()
    src.close()

def _fill_and_export_via_excel(template_abs_path, pdf_path, data):
    """
    Fills Excel and converts to PDF.
    Preserves line breaks in Remarks and handles XML corruption errors.
    """
    try:
        # data_only=False keeps formulas; keep_vba=False avoids many metadata crashes
        wb = openpyxl.load_workbook(template_abs_path, data_only=False, keep_vba=False)
        ws = wb.worksheets[0] 
        
        # --- CRITICAL FIX: Wipe sheet view properties to prevent 'NoneType' crashes ---
        ws.views.sheetView = [] 
    except Exception as e:
        raise RuntimeError(f"OpenPyXL Load Error: {str(e)}")

    # --- PAGE SETUP (Area A1:P79 with Centering) ---
    try:
        # Small margins help the FitToPage logic center the content better
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0, footer=0)
        ws.print_area = 'A1:P79' 
        
        if hasattr(ws, 'page_setup') and ws.page_setup is not None:
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 1
            ws.page_setup.fitToWidth = 1
            # AUTO CENTERING
            ws.page_setup.horizontalCentered = True
            ws.page_setup.verticalCentered = True
    except:
        pass

    def set_cell(addr, value, number_format=None):
        try:
            ws[addr] = value
            if number_format: ws[addr].number_format = number_format
        except: pass

    check = lambda condition: '/' if condition else ''
    
    cmf = data['cmf']; dates = data['dates']; formula_info = data['formula_info']
    color_req_obj = data['color_req_obj']; resins = data['resins']
    process_list = data['process_list']; spec_list = data['spec_list']
    final_prod_code = data['final_prod_code']

    # --- FILLING CELLS (USING EXACT COORDINATES PROVIDED) ---
    # --- GENERAL INFORMATION ---
    set_cell('F7', cmf.cm_no)
    set_cell('F9', formula_info.customer if formula_info else "")
    set_cell('F11', dates.form_made.strftime('%m/%d/%Y') if dates and dates.form_made else "")
    set_cell('F13', dates.date_required if dates else "")
    set_cell('F15', cmf.sm.name if cmf.sm else "")
    
    set_cell('F17', check(cmf.matching_type == 'new'))
    set_cell('I17', check(cmf.matching_type == 'rematch'))
    
    set_cell('F19', check(cmf.product_status == 'existing'))
    set_cell('L19', check(cmf.product_status == 'new'))
    
    set_cell('F21', formula_info.finished_product if formula_info else "")
    set_cell('F23', cmf.color_desc)

    # --- COLOR REQUIREMENT ---
    c_req_name = color_req_obj.name if color_req_obj else ""
    standard_reqs = ['transparent', 'opaque', 'translucent', 'metallic', 'fluorescent', 'pearlescent']
    req_map = {
        'transparent': 'F25', 'opaque': 'I25', 'translucent': 'L25', 
        'metallic': 'F27', 'fluorescent': 'I27', 'pearlescent': 'L27'
    }

    for addr in req_map.values(): set_cell(addr, "")
    set_cell('F29', "")
    set_cell('H29', "")

    if c_req_name in standard_reqs:
        set_cell(req_map[c_req_name], "/")
    elif c_req_name:
        set_cell('F29', "/")           # "Others" checkbox
        set_cell('H29', c_req_name)    # "Others" text value

    # --- SAMPLE COLORANT AVAILABLE ---
    set_cell('F31', check(cmf.is_sample_available is True))
    set_cell('I31', check(cmf.is_sample_available is False))

    # --- TYPE OF COLORANT ---
    set_cell('F33', check(cmf.colorant_type == 'MB'))
    set_cell('I33', check(cmf.colorant_type == 'DC'))
    is_other_colorant = cmf.colorant_type not in ('MB', 'DC')
    set_cell('L33', check(is_other_colorant))
    set_cell('O33', cmf.colorant_type if is_other_colorant else "")

    # --- DOSAGE, QTY ORDER, RESIN ---
    set_cell('F35', formula_info.dosage if formula_info else "")
    set_cell('F37', cmf.est_qty_order, number_format='#,##0.00 "KG"')
    set_cell('F39', resins)

    # --- PROCESS ---
    set_cell('F41', check('injection' in process_list))
    set_cell('I41', check('blow-molding' in process_list))
    set_cell('L41', check('film' in process_list))
    set_cell('F43', check('pipe-extrusion' in process_list))
    
    standard_procs = ['injection', 'blow-molding', 'film', 'pipe-extrusion']
    other_procs = [p for p in process_list if p not in standard_procs]
    set_cell('I43', check(bool(other_procs)))
    set_cell('K43', ", ".join(other_procs) if other_procs else "")

    # --- RESIN PROVIDED & MI ---
    set_cell('F45', cmf.qty_resin_testing)
    set_cell('F47', check(cmf.is_resin_provided is True))
    set_cell('I47', check(cmf.is_resin_provided is False))
    set_cell('F49', cmf.mi_c_resin)

    # --- COLOR GUIDE RETURN ---
    set_cell('F51', check(cmf.is_guide_to_return is True))
    set_cell('I51', check(cmf.is_guide_to_return is False))

    # --- OTHER SPECIFICATIONS ---
    set_cell('F53', check('Food Contact' in spec_list))
    set_cell('I53', check('Sunlight Exposure' in spec_list))
    
    standard_specs = ['Food Contact', 'Sunlight Exposure']
    other_specs = [s for s in spec_list if s not in standard_specs]
    set_cell('F55', check(bool(other_specs)))
    set_cell('H55', ", ".join(other_specs) if other_specs else "")

    # --- TEMPERATURE & LOW COST ---
    set_cell('F57', cmf.temperature)
    set_cell('F59', check(cmf.is_low_cost is True))
    set_cell('I59', check(cmf.is_low_cost is False))

    # --- REMARKS & PRODUCT CODE ---
    # Preserves line breaks from database and enables WrapText in Excel
    ws['C64'] = cmf.remarks
    ws['C64'].alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
    
    set_cell('D76', final_prod_code)

    # 1. SAVE EXCEL (With defensive check for to_tree crash)
    temp_dir = os.path.dirname(pdf_path)
    temp_xlsx = os.path.join(temp_dir, f"temp_{uuid.uuid4().hex}.xlsx")
    try:
        wb.save(temp_xlsx)
    except Exception:
        try:
            ws.page_setup = None # Strip page_setup if it causes 'to_tree' error
            wb.save(temp_xlsx)
        except Exception as e:
            raise RuntimeError(f"Excel Save Failure: {str(e)}")

    # 2. CONVERT TO PDF (Isolated Profile fixes "Refused to connect")
    profile_path = os.path.join(temp_dir, f"lo_profile_{uuid.uuid4().hex}")
    os.makedirs(profile_path, exist_ok=True)
    try:
        libreoffice_bin = _get_libreoffice_executable()
        profile_url = f"file:///{profile_path.replace(os.sep, '/')}"
        command = [
            f'"{libreoffice_bin}"', '--headless', f'"-env:UserInstallation={profile_url}"',
            '--convert-to pdf', f'--outdir "{temp_dir}"', f'"{temp_xlsx}"'
        ]
        subprocess.run(" ".join(command), shell=True, check=True, capture_output=True, timeout=30)
        generated_pdf = temp_xlsx.replace('.xlsx', '.pdf')
        if os.path.exists(generated_pdf):
            if os.path.exists(pdf_path): os.remove(pdf_path)
            os.rename(generated_pdf, pdf_path)
    finally:
        if os.path.exists(temp_xlsx): os.remove(temp_xlsx)
        shutil.rmtree(profile_path, ignore_errors=True)

def _fetch_cmf_data(cm_no):
    """Pulls everything needed from the DB."""
    cmf = tbl_cmf.objects.get(cm_no=cm_no)
    dates = tbl_cmf_dates.objects.filter(cm_no=cmf).first()
    formula_info = tbl_cmf_formula.objects.filter(cm_no=cmf).first()
    color_req_obj = tbl_cmf_color_req.objects.filter(cm_no=cmf).first()
    resins = ", ".join(list(tbl_resins_selected.objects.filter(cm_no=cmf).values_list('resin_no__abbreviation', flat=True)))
    process_list = list(tbl_cmf_process02.objects.filter(cmf_formula_no=formula_info).values_list('process_no__name', flat=True)) if formula_info else []
    spec_list = list(tbl_cmf_specification02.objects.filter(cm_no=cmf).values_list('spec_no__name', flat=True))
    final_prod_code = ""
    final_f = tbl_mb_extruder_formula.objects.filter(cm_no=cmf, is_final=True).select_related('code').first()
    if not final_f:
        final_f = tbl_dc_extruder_formula.objects.filter(cm_no=cmf, is_final=True).select_related('code').first()
    if final_f and final_f.code:
        final_prod_code = final_f.code.product_code
    return {
        'cmf': cmf, 'dates': dates, 'formula_info': formula_info,
        'color_req_obj': color_req_obj, 'resins': resins,
        'process_list': process_list, 'spec_list': spec_list,
        'final_prod_code': final_prod_code,
    }

@xframe_options_exempt
def print_cmf_preview(request, cm_no):
    """Django View to serve the inline PDF preview."""
    try:
        data = _fetch_cmf_data(cm_no)
    except Exception as e:
        return HttpResponseServerError(f"Data Fetch Error: {str(e)}")

    template_abs_path = os.path.abspath(TEMPLATE_PATH)
    with _excel_lock:
        with tempfile.TemporaryDirectory() as tmpdir:
            raw_pdf_path = os.path.join(tmpdir, "raw_output.pdf")
            final_pdf_path = os.path.join(tmpdir, "final_output.pdf")
            try:
                _fill_and_export_via_excel(template_abs_path, raw_pdf_path, data)
                # Resizing to exact 6.5 x 8.5 as requested
                _resize_pdf_to_fixed_size(raw_pdf_path, final_pdf_path, width_in=6.5, height_in=8.5)
                with open(final_pdf_path, 'rb') as f:
                    pdf_bytes = f.read()
            except Exception as e:
                return HttpResponseServerError(f"PDF System Error: {str(e)}")

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'inline'
    response['X-Frame-Options'] = 'SAMEORIGIN'
    return response

def log_cmf_print(request, cm_no):
    try:
        log_audit(request, "Printed", f"Printed CMF No: {cm_no}")
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)