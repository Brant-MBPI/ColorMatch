from datetime import datetime

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from ...models import tbl_formula01, tbl_formula02

def generate_formulation_excel(date_from=None, date_to=None):
    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulation Records"

    # Define Headers
    headers = [
        "Form ID", "Date", "Customer", "Product Code", 
        "Material Code", "Seq No", "Concentration", 
        "Total", "Remarks", "Is Deleted"
    ]
    
    # Write Headers with Bold Font
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = Font(bold=True)

    # Fetch Data
    queryset = tbl_formula01.objects.all()

    # Apply Date Filters if provided (mm/dd/yyyy format expected from frontend)
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    queryset = queryset.order_by('-form_id')

    row_index = 2
    for f01 in queryset:
        # Get ingredients for this header
        items = tbl_formula02.objects.filter(form=f01).order_by('sequence_no')
        
        # If there are no ingredients, still show the header info
        if not items.exists():
            data = [
                f01.form_id,
                f01.date.strftime('%m/%d/%Y') if f01.date else "",
                f01.customer,
                f01.prod_code,
                "", # Material Code
                "", # Seq No
                "", # Concentration
                f01.total_concentration,
                f01.notes,
                "Yes" if f01.is_deleted else "No"
            ]
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_index, column=col_num, value=value)
            row_index += 1
        else:
            # Create a row for every ingredient
            for f02 in items:
                data = [
                    f01.form_id,
                    f01.date.strftime('%m/%d/%Y') if f01.date else "",
                    f01.customer,
                    f01.prod_code,
                    f02.material_code,
                    f02.sequence_no,
                    f02.concentration, # Raw value, no formatting
                    f01.total_concentration,
                    f01.notes,
                    "Yes" if f01.is_deleted else "No"
                ]
                for col_num, value in enumerate(data, 1):
                    ws.cell(row=row_index, column=col_num, value=value)
                row_index += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    return wb

def export_formulation_excel(request):
    # Get params from URL
    date_from_raw = request.GET.get('date_from')
    date_to_raw = request.GET.get('date_to')
    
    date_from = None
    date_to = None

    # Parse dates if they exist
    try:
        if date_from_raw:
            date_from = datetime.strptime(date_from_raw, '%m/%d/%Y').date()
        if date_to_raw:
            date_to = datetime.strptime(date_to_raw, '%m/%d/%Y').date()
    except ValueError:
        pass # Fallback to no filter if date format is wrong

    # Generate the workbook
    wb = generate_formulation_excel(date_from, date_to)

    # Prepare Response
    filename = f"Formulation_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response