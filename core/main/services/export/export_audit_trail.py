import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.db.models import Q
from datetime import datetime
from main.models import tbl_audit_trail
from main.utils.log_audit_trail import log_audit # Ensure this is imported

def export_audit_trail_excel(request):
    # 1. Get Parameters
    dept_filter = request.GET.get('department', 'all')
    col_choice = request.GET.get('column_choice', 'all')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search_value = request.GET.get('search', '').strip()

    # 2. Build Queryset
    queryset = tbl_audit_trail.objects.select_related('user', 'user__role').all()

    # 3. Apply Filters
    if dept_filter != 'all':
        queryset = queryset.filter(user__role__department=dept_filter)

    if date_from and date_to:
        queryset = queryset.filter(timestamp__date__range=[date_from, date_to])

    if search_value:
        if col_choice == 'Username':
            queryset = queryset.filter(user__username__icontains=search_value)
        elif col_choice == 'Full Name':
            queryset = queryset.filter(Q(user__first_name__icontains=search_value) | Q(user__last_name__icontains=search_value))
        elif col_choice == 'Action':
            queryset = queryset.filter(action_type__icontains=search_value)
        elif col_choice == 'Email':
            queryset = queryset.filter(user__email__icontains=search_value)
        elif col_choice == 'Details':
            queryset = queryset.filter(details__icontains=search_value)
        else:
            queryset = queryset.filter(
                Q(user__username__icontains=search_value) | Q(user__first_name__icontains=search_value) |
                Q(user__last_name__icontains=search_value) | Q(action_type__icontains=search_value) |
                Q(details__icontains=search_value) | Q(user__email__icontains=search_value)
            )

    queryset = queryset.order_by('-timestamp')

    # 4. Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Trail"

    headers = ['Timestamp', 'Username', 'Full Name', 'Action', 'Details', 'Email', 'Department']
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True)

    for col_num, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 5. Fill Data
    for row_num, log in enumerate(queryset, 2):
        fname = log.user.first_name if log.user else ""
        lname = log.user.last_name if log.user else ""
        full_name = f"{lname}, {fname}" if lname else "---"
        
        ws.cell(row=row_num, column=1, value=log.timestamp.strftime('%m/%d/%Y %I:%M %p'))
        ws.cell(row=row_num, column=2, value=log.user.username if log.user else "System")
        ws.cell(row=row_num, column=3, value=full_name)
        ws.cell(row=row_num, column=4, value=log.action_type)
        ws.cell(row=row_num, column=5, value=log.details)
        ws.cell(row=row_num, column=6, value=log.user.email if log.user else "---")
        ws.cell(row=row_num, column=7, value=log.user.role.department if log.user and log.user.role else "---")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 65)

    # 6. Filename & Date Range Logic
    if date_from and date_to:
        try:
            d1 = datetime.strptime(date_from, '%Y-%m-%d').strftime('%m%d%y')
            d2 = datetime.strptime(date_to, '%Y-%m-%d').strftime('%m%d%y')
            file_range = f"{d1}-{d2}"
        except:
            file_range = datetime.now().strftime('%m%d%y')
    else:
        file_range = f"Full_{datetime.now().strftime('%m%d%y')}"

    filename = f"Audit_Trail_{file_range}.xlsx"

    # --- 7. LOG AUDIT ACTION ---
    log_details = f"Exported Audit Trail to Excel. Range: {date_from or 'All'} to {date_to or 'All'}."
    if dept_filter != 'all':
        log_details += f" Dept Filter: {dept_filter}."
    if search_value:
        log_details += f" Search Term: '{search_value}' (Col: {col_choice})."
    
    log_audit(request, "Exported", log_details)

    # 8. Return Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response