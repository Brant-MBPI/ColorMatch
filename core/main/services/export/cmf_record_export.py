import io
from datetime import datetime, date
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.db.models import Q
from main.utils.log_audit_trail import log_audit
from main.models import (
    tbl_cmf_dates, tbl_cmf_formula, tbl_cmf_pending_completed,
)

def get_export_data(date_from, date_to, include_completed, include_pending, include_rs):
    """
    Optimized retrieval. Filters by tbl_cmf_dates.form_made and 
    sorts results in ascending order.
    """
    def parse_date(d_str):
        if not d_str: return None
        try: return datetime.strptime(d_str.strip(), '%m/%d/%Y').date()
        except ValueError: return None

    date_from_parsed = parse_date(date_from)
    date_to_parsed = parse_date(date_to)

    # 1. Performance: Pre-fetch related info into dictionaries to avoid N+1 queries
    formula_map = {f.cm_no_id: f for f in tbl_cmf_formula.objects.all()}
    
    date_filter = Q()
    if date_from_parsed: date_filter &= Q(form_made__gte=date_from_parsed)
    if date_to_parsed: date_filter &= Q(form_made__lte=date_to_parsed)
    
    date_map_cmf = {d.cm_no_id: d for d in tbl_cmf_dates.objects.filter(date_filter, cm_no__isnull=False)}
    date_map_rs = {d.rs_no_id: d for d in tbl_cmf_dates.objects.filter(date_filter, rs_no__isnull=False)}

    pending_list = []
    completed_list = []

    # 2. Build Filter for status records
    status_filter = Q(cm_no__isnull=False)
    if include_rs:
        status_filter |= Q(rs_no__isnull=False)

    status_records = tbl_cmf_pending_completed.objects.filter(status_filter).select_related(
        'cm_no', 'cm_no__sm', 'rs_no', 'rs_no__sm_no', 'code'
    )

    for entry in status_records:
        is_cmf = True if entry.cm_no else False
        parent = entry.cm_no if is_cmf else entry.rs_no
        parent_key = parent.cm_no if is_cmf else parent.id
        
        # Pull date data from our pre-fetched maps
        dates = date_map_cmf.get(parent_key) if is_cmf else date_map_rs.get(parent_key)
        
        # If no date record found within the date range, skip this entry
        if not dates:
            continue

        formula = formula_map.get(parent_key) if is_cmf else None
        
        cust = formula.customer if formula else (parent.customer if not is_cmf else "---")
        prod_code = entry.code.product_code if entry.code else "---"
        raw_date = dates.form_made

        if entry.is_completed:
            if not include_completed: continue
            completed_list.append({
                "sort_date": raw_date or date(1900, 1, 1),
                "customer": cust,
                "code": prod_code,
                "date_request": raw_date.strftime('%m/%d/%Y') if raw_date else "---",
                "date_lab_received": dates.date_received_lab if dates else "---",
                "date_submitted": entry.date_submitted.strftime('%m/%d/%Y') if entry.date_submitted else "---",
                "ar_no": entry.ar_no or "---",
                "ar_date": entry.ar_date.strftime('%m/%d/%Y') if entry.ar_date else "---",
            })
        else:
            if not include_pending: continue
            sm_name = parent.sm.name if is_cmf and parent.sm else (parent.sm_no.name if not is_cmf and parent.sm_no else "---")
            end_prod = formula.finished_product if formula else (parent.finished_product if not is_cmf else "---")
            
            pending_list.append({
                "sort_date": raw_date or date(1900, 1, 1),
                "matching_no": parent.cm_no if is_cmf else parent.rs_no,
                "customer": cust,
                "date_form_made": raw_date.strftime('%m/%d/%Y') if raw_date else "---",
                "date_lab_received": dates.date_received_lab if dates else "---",
                "date_needed": dates.date_required if dates else "---",
                "target_date": f"due on {dates.due_date_lab.strftime('%m/%d/%y')}" if dates and dates.due_date_lab else "---",
                "end_product": end_prod,
                "color": parent.color_desc or "---",
                "matching_type": parent.matching_type or "---",
                "sm": sm_name,
                "reason": entry.reason or "pending",
            })

    # 3. Sort both lists ASCENDING by the creation date
    pending_list.sort(key=lambda x: x['sort_date'])
    completed_list.sort(key=lambda x: x['sort_date'])

    return pending_list, completed_list


def build_export_workbook(pending_rows, completed_rows, include_pending, include_completed):
    wb = Workbook()
    ws = wb.active
    ws.title = "CMF Records"

    # Define Header & Cell Styling
    header_font = Font(name="Arial", bold=True)
    normal_font = Font(name="Arial")
    # Gold, Accent 4, Lighter 40% (#FFD966)
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    row_idx = 1

    # --- Pending Section ---
    if include_pending and pending_rows:
        pending_headers = [
            "MATCHING No", "CUSTOMER", "Date Form Created",
            "Date Lab Received", "Date Needed by sales",
            "Lab Target Date", "Finished Product", "Color", 
            "MATCHING TYPE", "Salesman", "REASON"
        ]
        for col_idx, header in enumerate(pending_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        row_idx += 1

        for data in pending_rows:
            values = [
                data["matching_no"], data["customer"], data["date_form_made"],
                data["date_lab_received"], data["date_needed"], data["target_date"],
                data["end_product"], data["color"], data["matching_type"],
                data["sm"], data["reason"],
            ]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value).font = normal_font
            row_idx += 1

    # Spacer between Pending and Completed
    if include_pending and include_completed:
        row_idx += 3

    # --- Completed Section ---
    if include_completed and completed_rows:
        completed_headers = [
            "CUSTOMER", "CODE", "Date Form Created", "Date Lab Received",
            "DATE SUBMITTED", "AR#", "AR DATE"
        ]
        for col_idx, header in enumerate(completed_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        row_idx += 1

        for data in completed_rows:
            values = [
                data["customer"], data["code"], data["date_request"],
                data["date_lab_received"], data["date_submitted"],
                data["ar_no"], data["ar_date"],
            ]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value).font = normal_font
            row_idx += 1

    # --- AUTO-ADJUST COLUMN WIDTHS ---
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        
        # Set width with extra padding, capped for readability
        adjusted_width = max(12, min(max_length + 3, 50))
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def log_cmf_export_action(request):
    if request.method == 'POST':
        # Get the descriptive string sent from JS
        details = request.POST.get('details', 'Exported CMF Records from Preview')
        
        # Log it
        log_audit(request, "Exported", details)
        
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)