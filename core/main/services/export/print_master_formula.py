from datetime import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from ...models import tbl_master_formula, tbl_master_formula_info

def generate_master_formula_excel(date_from=None, date_to=None):
    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Formula Records"

    # 1. Define Headers
    headers = [
        "Form ID", "Date", "Customer", "Product Code", 
        "Material Code", "Seq No", "Concentration", 
        "Total", "Remarks", "Is Deleted"
    ]
    
    # Write Headers with Bold Font
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = Font(bold=True)

    # 2. Fetch Data
    queryset = tbl_master_formula.objects.all()

    # Apply Date Filters
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # Order by newest first
    queryset = queryset.order_by('-form_id')

    row_index = 2
    for mf in queryset:
        # Get ingredients for this master formula
        items = tbl_master_formula_info.objects.filter(form=mf).order_by('sequence_no')
        
        # Base data from the header (tbl_master_formula)
        # We format the date to mm/dd/yyyy here
        base_data = [
            mf.form_id,
            mf.date.strftime('%m/%d/%Y') if mf.date else "",
            mf.customer,
            mf.product_code,
        ]

        # If there are no ingredients, write one row with empty material info
        if not items.exists():
            data = base_data + ["", "", "", mf.total_concentration, mf.notes, "Yes" if mf.is_deleted else "No"]
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_index, column=col_num, value=value)
            row_index += 1
        else:
            # Create a row for every ingredient
            for item in items:
                data = base_data + [
                    item.material_code,
                    item.sequence_no,
                    item.concentration, # Raw numeric value
                    mf.total_concentration,
                    mf.notes,
                    "Yes" if mf.is_deleted else "No"
                ]
                for col_num, value in enumerate(data, 1):
                    ws.cell(row=row_index, column=col_num, value=value)
                row_index += 1

    # 3. Auto-adjust column widths for readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    return wb

def export_master_formula_excel(request):
    date_from_raw = request.GET.get('date_from')
    date_to_raw = request.GET.get('date_to')
    
    date_from = None
    date_to = None

    try:
        if date_from_raw:
            date_from = datetime.strptime(date_from_raw, '%m/%d/%Y').date()
        if date_to_raw:
            date_to = datetime.strptime(date_to_raw, '%m/%d/%Y').date()
    except ValueError:
        pass

    # --- DYNAMIC FILENAME LOGIC ---
    base_name = "Master_Formula_Export"
    
    if date_from and date_to:
        # e.g., Master_Formula_Export_01-01-2026_to_01-31-2026.xlsx
        filename = f"{base_name}_{date_from.strftime('%m-%d-%Y')}_to_{date_to.strftime('%m-%d-%Y')}.xlsx"
    elif date_from:
        # e.g., Master_Formula_Export_From_01-01-2026.xlsx
        filename = f"{base_name}_From_{date_from.strftime('%m-%d-%Y')}.xlsx"
    elif date_to:
        # e.g., Master_Formula_Export_Until_01-31-2026.xlsx
        filename = f"{base_name}_Until_{date_to.strftime('%m-%d-%Y')}.xlsx"
    else:
        # Exported All - use current date (e.g., Master_Formula_Export_August_19_2026.xlsx)
        filename = f"{base_name}_{datetime.now().strftime('%B_%d_%Y')}.xlsx"

    # Generate WB
    wb = generate_master_formula_excel(date_from, date_to)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response