import os
import tempfile
import threading
import uuid

import pythoncom
import win32com.client as win32
from django.contrib import messages
from django.http import HttpResponse, HttpResponseServerError
from django.shortcuts import redirect
from openpyxl import load_workbook

from main.models import (
    tbl_cmf, tbl_cmf_dates, tbl_cmf_formula, tbl_cmf_color_req,
    tbl_resins_selected, tbl_cmf_process02, tbl_cmf_specification02,
    tbl_mb_extruder_formula, tbl_dc_extruder_formula
)

# Excel COM automation isn't safe to run from multiple threads/requests at
# once. Serialize access so only one conversion happens at a time.
_excel_lock = threading.Lock()


def _build_cmf_workbook(cm_no):
    """Loads the template and writes DB data into it. Returns an openpyxl Workbook."""
    cmf = tbl_cmf.objects.get(cm_no=cm_no)
    dates = tbl_cmf_dates.objects.filter(cm_no=cmf).first()
    formula_info = tbl_cmf_formula.objects.filter(cm_no=cmf).first()
    color_req_obj = tbl_cmf_color_req.objects.filter(cm_no=cmf).first()

    resins = ", ".join(list(tbl_resins_selected.objects.filter(cm_no=cmf).values_list('resin_no__abbreviation', flat=True)))
    process_list = list(tbl_cmf_process02.objects.filter(cmf_formula_no=formula_info).values_list('process_no__name', flat=True)) if formula_info else []
    spec_list = list(tbl_cmf_specification02.objects.filter(cm_no=cmf).values_list('spec_no__name', flat=True))

    final_prod_code = ""
    final_f = tbl_mb_extruder_formula.objects.filter(cm_no=cmf, is_final=True).select_related('code').first()
    if not final_f:
        final_f = tbl_dc_extruder_formula.objects.filter(cm_no=cmf, is_final=True).select_related('code').first()
    if final_f and final_f.code:
        final_prod_code = final_f.code.product_code

    template_path = 'main/templates/print_excel/cmf_template.xlsx'
    wb = load_workbook(template_path)
    ws = wb.active

    ws['F6'] = cmf.cm_no
    ws['F7'] = formula_info.customer if formula_info else ""
    ws['F8'] = dates.form_made.strftime('%m/%d/%Y') if dates and dates.form_made else ""
    ws['F9'] = dates.date_required if dates else ""
    ws['F11'] = "✔" if cmf.matching_type == 'new' else ""
    ws['I11'] = "✔" if cmf.matching_type == 'rematch' else ""
    ws['F12'] = cmf.sm.name if cmf.sm else ""
    ws['F13'] = cmf.color_desc
    ws['F14'] = formula_info.finished_product if formula_info else ""

    c_req_name = color_req_obj.name if color_req_obj else ""
    standard_reqs = ['transparent', 'opaque', 'translucent', 'metallic', 'fluorescent', 'pearlescent']
    req_map = {'transparent': 'F17', 'opaque': 'I17', 'translucent': 'L17', 'metallic': 'F19', 'fluorescent': 'I19', 'pearlescent': 'L19'}
    if c_req_name in standard_reqs:
        ws[req_map[c_req_name]] = "✔"
    elif c_req_name:
        ws['F21'] = "✔"
        ws['H21'] = c_req_name

    ws['D21'] = resins
    ws['F24'] = "✔" if 'injection' in process_list else ""
    ws['I24'] = "✔" if 'blow-molding' in process_list else ""
    ws['M24'] = "✔" if 'film' in process_list else ""
    ws['F26'] = "✔" if 'pipe-extrusion' in process_list else ""

    standard_procs = ['injection', 'blow-molding', 'film', 'pipe-extrusion']
    other_procs = [p for p in process_list if p not in standard_procs]
    if other_procs:
        ws['I26'] = "✔"
        ws['L26'] = ", ".join(other_procs)

    ws['F28'] = cmf.qty_resin_testing
    ws['F29'] = "✔" if cmf.is_resin_provided is True else ""
    ws['I29'] = "✔" if cmf.is_resin_provided is False else ""
    ws['F30'] = cmf.mi_c_resin
    ws['F31'] = "✔" if cmf.is_sample_available is True else ""
    ws['I31'] = "✔" if cmf.is_sample_available is False else ""

    if cmf.colorant_type == 'MB':
        ws['F33'] = "✔"
    elif cmf.colorant_type == 'DC':
        ws['I33'] = "✔"
    else:
        ws['L33'] = "✔"
        ws['O33'] = cmf.colorant_type

    ws['F35'] = formula_info.dosage if formula_info else ""
    ws['F37'] = "✔" if cmf.is_guide_to_return is True else ""
    ws['I37'] = "✔" if cmf.is_guide_to_return is False else ""

    ws['F39'] = "✔" if 'Food Contact' in spec_list else ""
    ws['I39'] = "✔" if 'Sunlight Exposure' in spec_list else ""

    standard_specs = ['Food Contact', 'Sunlight Exposure']
    other_specs = [s for s in spec_list if s not in standard_specs]
    if other_specs:
        ws['F41'] = "✔"
        ws['G41'] = ", ".join(other_specs)

    ws['F43'] = cmf.temperature
    ws['F44'] = "✔" if cmf.is_low_cost is True else ""
    ws['I44'] = "✔" if cmf.is_low_cost is False else ""

    ws['C48'] = cmf.remarks
    ws['D62'] = final_prod_code

    return wb


def _convert_xlsx_to_pdf_via_excel(xlsx_path, pdf_path):
    """
    Drives a real Excel install through COM to export the sheet as PDF.
    Must run inside pythoncom.CoInitialize()/CoUninitialize() on whatever
    thread calls it, and must guarantee Excel.Quit() even on failure so
    no orphaned EXCEL.EXE processes pile up on the server.
    """
    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        # DispatchEx starts a fresh, isolated Excel instance rather than
        # attaching to one that might already be open/in use.
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(xlsx_path)
        ws = wb.Worksheets(1)

        # 0 = xlTypePDF
        ws.ExportAsFixedFormat(0, pdf_path)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def print_cmf_preview(request, cm_no):
    """
    Builds the workbook from the (reusable) template, converts it to PDF
    via Excel COM automation for inline browser preview, and cleans up
    every temp file before returning — nothing is left on disk waiting
    for the client to close anything.
    """
    try:
        wb = _build_cmf_workbook(cm_no)
    except tbl_cmf.DoesNotExist:
        messages.error(request, f"Error: CMF No. '{cm_no}' was not found.")
        return redirect('cmf_entry')
    except Exception as e:
        messages.error(request, f"System Error: {str(e)}")
        return redirect('cmf_entry')

    with tempfile.TemporaryDirectory() as tmpdir:
        file_id = uuid.uuid4().hex
        xlsx_path = os.path.join(tmpdir, f"{file_id}.xlsx")
        pdf_path = os.path.join(tmpdir, f"{file_id}.pdf")
        wb.save(xlsx_path)

        try:
            with _excel_lock:
                _convert_xlsx_to_pdf_via_excel(xlsx_path, pdf_path)
        except Exception as e:
            return HttpResponseServerError(f"PDF conversion failed: {str(e)}")

        if not os.path.exists(pdf_path):
            return HttpResponseServerError("PDF conversion failed: no output file produced.")

        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()
    # TemporaryDirectory context manager deletes xlsx + pdf here, unconditionally.

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    # inline (no filename) keeps it in the browser's PDF viewer instead of
    # prompting a "Save As" with a suggested name.
    response['Content-Disposition'] = 'inline'
    return response