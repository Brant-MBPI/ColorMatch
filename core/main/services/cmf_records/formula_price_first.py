import io
import os
import re
import json
import tempfile
import uuid
from decimal import Decimal
from datetime import datetime, date

# Cross-platform Excel libraries
import openpyxl
from openpyxl.styles import Font, Alignment
import msoffcrypto

from django.db.models import Max
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from main.models import (
    tbl_cmf, tbl_cmf_formula, tbl_dc_extruder_formula, 
    tbl_dc_extruder_materials, tbl_dc_extruder_version, 
    tbl_mb_extruder_formula, tbl_mb_extruder_formula02, 
    tbl_resins_selected
)

# Configuration
FORMULA_TEMPLATE_PATH = os.path.join('main', 'templates', 'print_excel', 'Formula.xlsx')
FORMULA_TEMPLATE_PASSWORD = "maranatha101"

COLUMN_ORDER = [
    'date', 'customer', 'classification', 'prod_code', 'resin', 'mat_code',
    'mat_conc', 'end_product', 'total', 'others', 'dosage', 'salesman_name',
    'cmf_no', 'html', 'c', 'm', 'y', 'k', 'remarks'
]
DATA_START_ROW = 2 

def _autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length: max_length = length
            except: pass
        ws.column_dimensions[column].width = max_length + 3

def _fill_formula_sheet_logic(template_abs_path, rows):
    # --- 1. DECRYPT TEMPLATE ---
    decrypted_tmp = io.BytesIO()
    with open(template_abs_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=FORMULA_TEMPLATE_PASSWORD)
        office_file.decrypt(decrypted_tmp)

    # --- 2. LOAD & STYLE WORKBOOK ---
    wb = openpyxl.load_workbook(decrypted_tmp)
    ws = wb["Formula"] if "Formula" in wb.sheetnames else wb.active
    
    # Define Fonts
    header_font = Font(name='Arial', size=14, bold=True)
    content_font = Font(name='Arial', size=12, bold=True)
    
    # Define Alignments based on your screenshot requirements
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Mapping columns to specific alignments to ensure consistency
    # (Based on standard Excel report layouts)
    COLUMN_ALIGNMENTS = {
        'date': align_left,
        'customer': align_left,
        'classification': align_center,
        'prod_code': align_left,
        'resin': align_left,
        'mat_code': align_center,
        'mat_conc': align_right, # Numbers
        'end_product': align_left,
        'total': align_right,    # Numbers
        'others': align_left,
        'dosage': align_center,
        'salesman_name': align_left,
        'cmf_no': align_center,
        'html': align_center,
        'c': align_center,
        'm': align_center,
        'y': align_center,
        'k': align_center,
        'remarks': align_left,
    }

    # Apply Header Styling (Row 1)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = align_center # Headers always centered

    # --- 3. FILL CONTENT ROWS ---
    for r_idx, row_data in enumerate(rows):
        excel_row = DATA_START_ROW + r_idx
        for c_idx, field in enumerate(COLUMN_ORDER):
            cell = ws.cell(row=excel_row, column=c_idx + 1)
            val = row_data.get(field, "")
            
            # IMPROVED TYPE HANDLING: 
            # Try to keep numbers as floats so Excel treats them as numbers
            if val == "no data" or val == "None" or val == "":
                cell.value = str(val)
            else:
                try:
                    # If it's a number, convert to float for Excel
                    cell.value = float(val)
                except (ValueError, TypeError):
                    cell.value = str(val)
            
            # Apply Content Styling
            cell.font = content_font
            
            # APPLY EXPLICIT ALIGNMENT from our map
            cell.alignment = COLUMN_ALIGNMENTS.get(field, align_left)

    _autofit_columns(ws)

    # --- 4. SAVE UNPROTECTED TO BUFFER ---
    unprotected_buffer = io.BytesIO()
    wb.save(unprotected_buffer)
    unprotected_buffer.seek(0)

    # --- 5. RE-ENCRYPT FOR OUTPUT ---
    encrypted_buffer = io.BytesIO()
    file_to_encrypt = msoffcrypto.OfficeFile(unprotected_buffer)
    file_to_encrypt.encrypt(FORMULA_TEMPLATE_PASSWORD, encrypted_buffer)
    
    return encrypted_buffer.getvalue()

# ... (Keep _build_price_first_row_list, export_formula_by_date, 
#      and download_price_first_excel exactly as they were) ...

def _build_price_first_row_list(items):
    """Shared helper to build the list of data dictionaries from models."""
    results = []
    for item in items:
        f_id = item['id']
        f_type = item['type']
        
        if f_type == 'MB':
            header = tbl_mb_extruder_formula.objects.select_related('code', 'cm_no', 'rs_no').get(pk=f_id)
            qs = tbl_mb_extruder_formula02.objects.filter(mb=header).order_by('id')
            ingredients_list = [{'material': ing.material, 'value': float(ing.value or 0)} for ing in qs]
        else:
            header = tbl_dc_extruder_formula.objects.select_related('code', 'cm_no', 'rs_no').get(pk=f_id)
            max_v = tbl_dc_extruder_version.objects.filter(material__dc=header).aggregate(Max('version_no'))['version_no__max']
            ingredients_list = []
            if max_v:
                version_data = tbl_dc_extruder_version.objects.filter(material__dc=header, version_no=max_v).select_related('material')
                ingredients_list = [{'material': v.material.material, 'value': float(v.value or 0)} for v in version_data]

        customer, dosage, end_product, salesman, matching_type = "", 0, "", "", ""
        
        if header.cm_no:
            formula_info = tbl_cmf_formula.objects.filter(cm_no=header.cm_no).first()
            if formula_info:
                customer, dosage, end_product = formula_info.customer, formula_info.dosage, formula_info.finished_product
            salesman = header.cm_no.sm.name if header.cm_no.sm else ""
            matching_type = header.cm_no.matching_type
        elif header.rs_no:
            customer, dosage, end_product = header.rs_no.customer, header.rs_no.dosage, header.rs_no.finished_product
            salesman = header.rs_no.sm_no.name if header.rs_no.sm_no else ""
            matching_type = header.rs_no.matching_type

        resins_list = list(tbl_resins_selected.objects.filter(
            cm_no=header.cm_no if header.cm_no else None,
            rs_no=header.rs_no if header.rs_no else None
        ).values_list('resin_no__abbreviation', flat=True))
        
        resin_str = ", ".join(resins_list)

        others_val = "new matching"
        if matching_type == 'rematch' and header.cm_no:
            curr_cm = header.cm_no.cm_no
            match = re.match(r"([A-Z0-9]+)([a-z]+)", curr_cm)
            if match:
                base_code = match.group(1)
                prev_cmf = tbl_cmf.objects.filter(cm_no__startswith=base_code).exclude(cm_no=curr_cm).order_by('-cm_no').first()
                if prev_cmf:
                    pc_check = tbl_mb_extruder_formula.objects.filter(cm_no=prev_cmf, is_final=True).select_related('code').first() or \
                               tbl_dc_extruder_formula.objects.filter(cm_no=prev_cmf, is_final=True).select_related('code').first()
                    others_val = f"rematch of {pc_check.code.product_code if pc_check and pc_check.code else 'Unknown'}"
        elif matching_type == 'request': others_val = "request"

        total_conc = sum([i['value'] for i in ingredients_list])
        
        for ing in ingredients_list:
            results.append({
                'date': header.date.strftime('%B %d, %Y') if header.date else "no data",
                'customer': customer,
                'classification': f_type.lower(),
                'prod_code': header.code.product_code if header.code else "no data",
                'resin': resin_str,
                'mat_code': ing['material'],
                'mat_conc': f"{ing['value']:.6f}",
                'end_product': end_product,
                'total': f"{total_conc:.2f}",
                'others': others_val,
                'dosage': f"{float(dosage or 0):.2f}",
                'salesman_name': salesman,
                'cmf_no': header.cm_no.cm_no if header.cm_no else (header.rs_no.rs_no if header.rs_no else "none"),
                'html': (header.html or "no data").replace('#', ''),
                'c': int(header.c) if header.c is not None else "no data",
                'm': int(header.m) if header.m is not None else "no data",
                'y': int(header.y) if header.y is not None else "no data",
                'k': int(header.k) if header.k is not None else "no data",
                'remarks': ''
            })
    return results

def get_price_first_data(request):
    try:
        items = json.loads(request.GET.get('items', '[]'))
        data = _build_price_first_row_list(items)
        return JsonResponse({'data': data})
    except Exception as e:
        return HttpResponseBadRequest(str(e))

def export_formula_by_date(request):
    date_from_str = request.GET.get('from')
    date_to_str = request.GET.get('to')
    try:
        date_from = datetime.strptime(date_from_str, '%m/%d/%Y').date()
        date_to = datetime.strptime(date_to_str, '%m/%d/%Y').date()
        mb_ids = tbl_mb_extruder_formula.objects.filter(date__range=[date_from, date_to]).values_list('mb_no', flat=True)
        dc_ids = tbl_dc_extruder_formula.objects.filter(date__range=[date_from, date_to]).values_list('dc_no', flat=True)
        items = [{'type': 'MB', 'id': i} for i in mb_ids] + [{'type': 'DC', 'id': i} for i in dc_ids]
        if not items: return HttpResponseBadRequest("No records found.")
        row_dicts = _build_price_first_row_list(items)
        template_abs_path = os.path.abspath(FORMULA_TEMPLATE_PATH)
        file_bytes = _fill_formula_sheet_logic(template_abs_path, row_dicts)
        response = HttpResponse(file_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Formula_Export_{date_from_str.replace("/","-")}.xlsx"'
        return response
    except Exception as e:
        return HttpResponseBadRequest(f"Export failed: {str(e)}")

def download_price_first_excel(request):
    if request.method != 'POST': return HttpResponseBadRequest("POST required.")
    try:
        payload = json.loads(request.body)
        rows = payload.get('rows', [])
    except: return HttpResponseBadRequest("Invalid JSON.")
    if not rows: return HttpResponseBadRequest("No rows.")
    row_dicts = []
    for row in rows:
        row_dicts.append({field: (row[i] if i < len(row) else "") for i, field in enumerate(COLUMN_ORDER)})
    template_abs_path = os.path.abspath(FORMULA_TEMPLATE_PATH)
    try:
        file_bytes = _fill_formula_sheet_logic(template_abs_path, row_dicts)
    except Exception as e:
        return HttpResponseBadRequest(f"Excel export failed: {str(e)}")
    response = HttpResponse(file_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Formula.xlsx"'
    return response